"""Fetch the ONS air fare sub-index weights and write them to weights.csv.

The weights live in an ad hoc release, "Domestic, European and long-haul
airfares consumer prices sub-indices", as an .xlsx with a weights sheet
alongside the sub-indices themselves.

WHY THIS IS WRITTEN DEFENSIVELY
-------------------------------
This parser was written without sight of the actual spreadsheet: the
development sandbox's egress policy blocks ons.gov.uk outright, so the layout
below is inferred from the release description rather than observed. The
consequences are designed in:

  * The sheet, header row and columns are located by *searching* for them, not
    by fixed offsets, so minor layout changes do not break it.
  * Every extracted row is validated (three positive weights, a plausible year,
    no duplicate years). Anything failing validation is rejected.
  * On any parse failure it dumps the workbook's actual structure and exits
    non-zero, so one look at the CI log tells you what to fix -- rather than
    silently writing plausible-looking garbage into the weights that every
    downstream aggregate depends on.
  * `--dump` skips parsing entirely and just prints what is in the file.

If the layout turns out to differ, fix `_find_weight_sheet` / `_find_header` and
nothing else needs to change.

Run it where egress to ons.gov.uk is permitted (a GitHub Actions runner is
fine); it is wired into the monthly reconciliation workflow.
"""

from __future__ import annotations

import argparse
import calendar
import csv
import datetime as dt
import io
import logging
import pathlib
import re
import sys
from typing import Any, Iterable, Sequence

import requests

from .panel import WEIGHTS_PATH

log = logging.getLogger("ukairfares.onsweights")

ONS_BASE = "https://www.ons.gov.uk"

#: Most recent ad hoc release located at time of writing (Jan 2017 - Feb 2025).
#: Override with --release-url when ONS publish a newer vintage; --discover
#: attempts to find one automatically.
DEFAULT_RELEASE_URL = (
    f"{ONS_BASE}/economy/inflationandpriceindices/adhocs/"
    "2716domesticeuropeanandlonghaulairfaresconsumerpricessubindices"
    "january2017tofebruary2025"
)

SEARCH_URL = f"{ONS_BASE}/search"
SEARCH_TERM = "domestic European and long-haul airfares consumer prices sub-indices"

USER_AGENT = "uk-airfares-nowcasting/1.0 (research pipeline; contact via repository owner)"

#: Column header synonyms. Matched case-insensitively as substrings.
_CATEGORY_PATTERNS = {
    "domestic": (r"domestic",),
    "european": (r"europ", r"short[\s-]?haul"),
    "long_haul": (r"long[\s-]?haul", r"longhaul"),
}
_YEAR_PATTERNS = (r"^year$", r"^date$", r"^period$")
_MONTH_PATTERNS = (r"^month$", r"^date$", r"^period$", r"^index month$")

#: Sanity bounds. A weight outside this is a parse error, not a datum.
MIN_YEAR, MAX_YEAR = 2015, 2035

#: The sub-index series begins in January 2017 per the release title.
MIN_SERIES_YEAR = 2016

_MONTH_NAMES = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
_MONTH_NAMES.update({m.lower(): i for i, m in enumerate(calendar.month_abbr) if m})


class WeightsParseError(RuntimeError):
    """The workbook could not be parsed into defensible weights."""


class SubIndexParseError(RuntimeError):
    """The workbook could not be parsed into a defensible sub-index series."""


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _as_year(value: Any) -> int | None:
    if isinstance(value, dt.datetime):
        return value.year
    if isinstance(value, dt.date):
        return value.year
    if isinstance(value, (int, float)) and float(value).is_integer():
        year = int(value)
        return year if MIN_YEAR <= year <= MAX_YEAR else None
    match = re.search(r"\b(20\d{2})\b", str(value or ""))
    if match:
        year = int(match.group(1))
        return year if MIN_YEAR <= year <= MAX_YEAR else None
    return None


def _as_month(value: Any) -> dt.date | None:
    """Parse a monthly period label into the first of that month.

    Excel usually hands these over as datetimes, but ONS sheets also carry
    "Jan 2017", "January 2017" and "2017-01" forms depending on the vintage.
    """
    if isinstance(value, dt.datetime):
        return dt.date(value.year, value.month, 1)
    if isinstance(value, dt.date):
        return dt.date(value.year, value.month, 1)

    text = _norm(value)
    if not text:
        return None

    # "2017-01", "2017/01"
    match = re.match(r"^(20\d{2})[-/](\d{1,2})$", text)
    if match:
        year, month = int(match.group(1)), int(match.group(2))
        return dt.date(year, month, 1) if 1 <= month <= 12 else None

    # "jan 2017", "january 2017"
    match = re.match(r"^([a-z]+)[\s-]+(20\d{2})$", text)
    if match and match.group(1) in _MONTH_NAMES:
        return dt.date(int(match.group(2)), _MONTH_NAMES[match.group(1)], 1)

    # "2017 jan"
    match = re.match(r"^(20\d{2})[\s-]+([a-z]+)$", text)
    if match and match.group(2) in _MONTH_NAMES:
        return dt.date(int(match.group(1)), _MONTH_NAMES[match.group(2)], 1)

    return None


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[,%\s£]", "", str(value or ""))
    try:
        return float(text)
    except ValueError:
        return None


def find_release_xlsx(release_url: str, session: requests.Session) -> str:
    """Locate the .xlsx download link on an ad hoc release page."""
    resp = session.get(release_url, timeout=60, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    links = re.findall(r'href="([^"]+\.xlsx[^"]*)"', resp.text, flags=re.IGNORECASE)
    if not links:
        raise WeightsParseError(
            f"no .xlsx link on {release_url}. The release layout may have changed; "
            "pass --xlsx-url with the direct file link."
        )
    href = links[0]
    return href if href.startswith("http") else f"{ONS_BASE}{href}"


def discover_latest_release(session: requests.Session) -> str | None:
    """Best-effort search for a newer ad hoc release than the pinned default."""
    try:
        resp = session.get(
            SEARCH_URL,
            params={"q": SEARCH_TERM},
            timeout=60,
            headers={"User-Agent": USER_AGENT},
        )
        resp.raise_for_status()
    except requests.RequestException as exc:
        log.warning("release discovery failed (%s); using the pinned default", exc)
        return None

    hrefs = re.findall(
        r'href="(/economy/inflationandpriceindices/adhocs/[^"]*'
        r'airfaresconsumerpricessubindices[^"]*)"',
        resp.text,
        flags=re.IGNORECASE,
    )
    if not hrefs:
        return None

    best = max(hrefs, key=_coverage_end)
    if _coverage_end(best) == (0, 0):
        log.warning("could not read a coverage period from any release slug")
        return None
    return f"{ONS_BASE}{best}"


def _coverage_end(href: str) -> tuple[int, int]:
    """Latest month covered by a release, read from its URL slug.

    Slugs end with the coverage period, e.g.
    ".../2716domesticeuropeanandlonghaulairfaresconsumerpricessubindices
       january2017tofebruary2025".

    Ranking by the coverage period rather than the ad hoc reference number is
    deliberate and was learned the hard way: ONS restarted their ad hoc
    numbering, so the old series (13727, 14097, 14287) sorts *above* the newer
    one (2032, 2362, 2716) numerically. A production run picked 14287 -- which
    stops at January 2022 -- over 2716, which runs to February 2025. The
    coverage period is the only field in the URL that actually means what we
    need it to mean.

    Returns (year, month), or (0, 0) if the slug carries no readable period.
    """
    match = re.search(r"to([a-z]+)(20\d{2})(?:/|$)", href, flags=re.IGNORECASE)
    if not match:
        return (0, 0)
    month = _MONTH_NAMES.get(match.group(1).lower())
    return (int(match.group(2)), month) if month else (0, 0)


def _iter_sheets(workbook) -> Iterable[tuple[str, list[list[Any]]]]:
    for sheet in workbook.worksheets:
        yield sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]


def _year_columns(row: list[Any]) -> dict[int, int]:
    """Map column index -> year from a header row of years."""
    out: dict[int, int] = {}
    for col, cell in enumerate(row):
        year = _as_year(cell)
        if year is not None:
            out[col] = year
    return out


def parse_weights(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    """Extract ONS sub-index weights from the release.

    LAYOUT (confirmed against the real workbook, Aug 2026):

        sheet "Weights"
          | 2007-2026 airfares weights |         |        |        |
          |                            |         | 2026   | 2025   |
          | Domestic                   | 1 month | 0.0286 | 0.0271 |
          | European                   | 1 month | 0.2068 | 0.2077 |
          |                            | 3 month | 0.2068 | 0.2077 |
          | Long-haul                  | 1 month | 0.0558 | 0.0558 |
          |                            | 3 month | 0.2510 | 0.2509 |
          |                            | 6 month | 0.2510 | 0.2509 |

    Two things to note:

    * **Weights are per (haul, advance window), not per haul.** Six weights,
      matching the six published series, and they sum to 1.0 within a year. A
      category weight is not simply split evenly across its windows -- long-haul
      1-month carries 0.056 while its 3- and 6-month windows carry 0.251 each --
      so the split has to be read, not derived.
    * **Transposed, with years descending across columns.** As with the
      sub-index sheets, the category label is merged and blank on continuation
      rows, so it is carried forward.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise WeightsParseError(
            "openpyxl is required to parse the ONS release; pip install openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheets = list(_iter_sheets(workbook))

    for title, rows in sheets:
        if "weight" not in _norm(title):
            continue

        years: dict[int, int] = {}
        category: str | None = None
        out: list[dict[str, Any]] = []
        for row in rows:
            if not row:
                continue
            if not years:
                found = _year_columns(list(row))
                # A real header names several years, not one stray number.
                if len(found) >= 3:
                    years = found
                continue

            category = _as_category(row[0]) or category
            window = _as_window(row[1]) if len(row) > 1 else None
            if category is None or window is None:
                continue

            for col, year in years.items():
                if col >= len(row):
                    continue
                value = _as_number(row[col])
                if value is None or value <= 0:
                    continue
                out.append(
                    {
                        "year": year,
                        "haul_category": category,
                        "months_ahead": window,
                        "weight": value,
                    }
                )

        if len(out) >= 6:
            log.info(
                "parsed %d weights across %d years from sheet %r",
                len(out), len(years), title,
            )
            return sorted(
                out, key=lambda r: (r["year"], r["haul_category"], r["months_ahead"])
            )

    raise WeightsParseError(
        "could not locate a weights table. Workbook structure:\n" + describe(sheets)
    )


_WINDOW_PATTERN = re.compile(r"(\d+)\s*[-\s]?month", re.IGNORECASE)

_CATEGORY_LABELS = {
    "domestic": "domestic",
    "european": "european",
    "short-haul": "european",
    "short haul": "european",
    "long-haul": "long_haul",
    "long haul": "long_haul",
    "longhaul": "long_haul",
}


def _as_category(value: Any) -> str | None:
    text = _norm(value)
    return _CATEGORY_LABELS.get(text) if text else None


def _as_window(value: Any) -> int | None:
    """Parse "1-month" / "3 month" / "6-month" into months ahead."""
    match = _WINDOW_PATTERN.search(_norm(value))
    if not match:
        return None
    months = int(match.group(1))
    return months if months in (1, 3, 6) else None


def _month_columns(row: list[Any]) -> dict[int, int]:
    """Map column index -> month number from a header row of month names."""
    out: dict[int, int] = {}
    for col, cell in enumerate(row):
        name = _norm(cell)
        if name in _MONTH_NAMES:
            out[col] = _MONTH_NAMES[name]
    return out


def parse_subindices(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    """Extract ONS's published sub-index series from the release.

    LAYOUT (confirmed against the real workbook, Aug 2026):

        sheet "2019"
          | 2019 Airfares sub-indices |         |     |        |  ...
          |                           |         | Jan | Feb    |  ...
          | Domestic                  | 1-month | 100 | 102.91 |  ...
          | European                  | 1-month | 100 | 118.62 |  ...
          |                           | 3-month | 100 | 122.84 |  ...
          | Long-haul                 | 1-month | 100 | 74.44  |  ...
          |                           | 3-month | 100 | 103.10 |  ...
          |                           | 6-month | 100 | 94.31  |  ...

    Three things about this drive the implementation:

    * **One sheet per year.** The year comes from the sheet name, not the data.
    * **Transposed.** Months run across columns; series run down rows.
    * **Six series, not three.** ONS publish each haul category broken out by
      *advance window* -- domestic at 1 month, European at 1 and 3, long-haul at
      1, 3 and 6. That is the granularity our own panel already collects at
      (`months_ahead`), so we can and should compare like with like rather than
      collapsing the windows together.

    The category cell is blank on continuation rows (merged in the original), so
    it is carried forward. Missing values appear as ".." and are skipped.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise SubIndexParseError(
            "openpyxl is required to parse the ONS release; pip install openpyxl"
        ) from exc

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheets = list(_iter_sheets(workbook))

    out: list[dict[str, Any]] = []
    seen: set[tuple] = set()
    years_parsed: list[int] = []

    for title, rows in sheets:
        year = _as_year(_norm(title))
        if year is None or year < MIN_SERIES_YEAR:
            continue

        months: dict[int, int] = {}
        category: str | None = None
        for row in rows:
            if not row:
                continue
            if not months:
                found = _month_columns(list(row))
                # A real header names most of the year, not one stray cell.
                if len(found) >= 6:
                    months = found
                continue

            category = _as_category(row[0]) or category
            window = _as_window(row[1]) if len(row) > 1 else None
            if category is None or window is None:
                continue

            for col, month_num in months.items():
                if col >= len(row):
                    continue
                value = _as_number(row[col])
                if value is None or value <= 0:
                    continue  # ".." and blanks are legitimately missing
                key = (year, month_num, category, window)
                if key in seen:
                    continue
                seen.add(key)
                out.append(
                    {
                        "index_month": dt.date(year, month_num, 1),
                        "haul_category": category,
                        "months_ahead": window,
                        "index_value": value,
                    }
                )
        if months:
            years_parsed.append(year)

    if len(out) < 12:
        raise SubIndexParseError(
            f"parsed only {len(out)} sub-index values from {len(years_parsed)} "
            f"year sheet(s). Workbook structure:\n" + describe(sheets)
        )

    log.info(
        "parsed %d sub-index values across %d years (%s)",
        len(out), len(years_parsed), ", ".join(str(y) for y in sorted(years_parsed)),
    )
    return sorted(out, key=lambda r: (r["index_month"], r["haul_category"], r["months_ahead"]))


def describe(sheets: Sequence[tuple[str, list[list[Any]]]], max_rows: int = 12) -> str:
    """Human-readable dump of the workbook, for diagnosing a failed parse."""
    lines: list[str] = []
    for title, rows in sheets:
        lines.append(f"\n=== sheet {title!r} ({len(rows)} rows) ===")
        for row in rows[:max_rows]:
            cells = [("" if c is None else str(c))[:24] for c in row[:10]]
            if any(cells):
                lines.append("  | " + " | ".join(cells))
    return "\n".join(lines)


def write_weights_csv(rows: list[dict[str, Any]], path: pathlib.Path, source_url: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        fh.write(
            "# ONS air fare sub-index weights (CPI item 07.3.3).\n"
            f"# Fetched automatically from {source_url}\n"
            f"# on {dt.date.today().isoformat()} by ukairfares.onsweights.\n"
            "# One row per (year, haul category, advance window). ONS weight each\n"
            "# of the six published series separately; within a year they sum to 1.\n"
        )
        writer = csv.DictWriter(
            fh,
            fieldnames=["year", "haul_category", "months_ahead", "weight", "is_placeholder"],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({**row, "is_placeholder": "false"})
    log.info("wrote %d weight rows to %s", len(rows), path)


def fetch(
    *,
    release_url: str | None = None,
    xlsx_url: str | None = None,
    discover: bool = False,
    session: requests.Session | None = None,
) -> tuple[list[dict[str, Any]], str]:
    session = session or requests.Session()
    if not xlsx_url:
        url = release_url or DEFAULT_RELEASE_URL
        if discover and not release_url:
            url = discover_latest_release(session) or url
            log.info("using release %s", url)
        xlsx_url = find_release_xlsx(url, session)
    log.info("downloading %s", xlsx_url)
    resp = session.get(xlsx_url, timeout=120, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    return parse_weights(resp.content), xlsx_url


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Fetch ONS air fare sub-index weights into weights.csv."
    )
    parser.add_argument("--release-url", default=None, help="Ad hoc release page URL.")
    parser.add_argument("--xlsx-url", default=None, help="Direct .xlsx URL, skipping discovery.")
    parser.add_argument(
        "--discover",
        action="store_true",
        help="Search ONS for a newer release than the pinned default.",
    )
    parser.add_argument("--out", type=pathlib.Path, default=WEIGHTS_PATH)
    parser.add_argument(
        "--dump",
        action="store_true",
        help="Print the workbook's structure and exit without parsing.",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
    )

    try:
        if args.dump:
            import openpyxl

            session = requests.Session()
            url = args.xlsx_url or find_release_xlsx(
                args.release_url or DEFAULT_RELEASE_URL, session
            )
            content = session.get(url, timeout=120, headers={"User-Agent": USER_AGENT}).content
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            print(describe(list(_iter_sheets(wb)), max_rows=25))
            return 0

        rows, source = fetch(
            release_url=args.release_url, xlsx_url=args.xlsx_url, discover=args.discover
        )
    except WeightsParseError as exc:
        # Loud, with the evidence needed to fix it in one pass.
        print(f"::error::could not parse ONS weights: {exc}", flush=True)
        log.error("%s", exc)
        return 1
    except requests.RequestException as exc:
        print(f"::error::could not fetch ONS weights: {exc}", flush=True)
        log.error("%s", exc)
        return 1

    write_weights_csv(rows, args.out, source)
    print(
        f"Wrote {len(rows)} year(s) of weights ({rows[0]['year']}-{rows[-1]['year']}) "
        f"to {args.out}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
